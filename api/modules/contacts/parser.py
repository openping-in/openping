import csv
import io
from collections.abc import Iterator
from pathlib import Path
from typing import BinaryIO

import xlrd
from openpyxl import load_workbook

_HEADER_ALIASES = {
    "name": {"name"},
    "phone_country_code": {"phone_country_code", "country_code"},
    "phone_number": {"phone_number", "phone", "number"},
    "instagram_username": {"instagram_username", "instagram"},
    "facebook_user_id": {"facebook_user_id", "facebook"},
}


def _normalize_header(header: str) -> str:
    return header.strip().lower().replace(" ", "_")


def _build_column_map(headers: list[str]) -> dict[str, str]:
    normalized = {_normalize_header(header): header for header in headers}
    column_map: dict[str, str] = {}

    for field, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                column_map[field] = normalized[alias]
                break

    missing_fields = {"name", "phone_country_code", "phone_number"} - column_map.keys()
    if missing_fields:
        missing = ", ".join(sorted(missing_fields))
        raise ValueError(f"Missing required columns: {missing}")

    return column_map


def _cell_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _dict_row_to_contact_data(
    row: dict[str, str | None], column_map: dict[str, str]
) -> dict | None:
    def get(field: str) -> str:
        column = column_map.get(field)
        if column is None:
            return ""
        return _cell_value(row.get(column))

    name = get("name")
    phone_country_code = get("phone_country_code")
    phone_number = get("phone_number")
    if not name or not phone_country_code or not phone_number:
        return None

    return {
        "name": name,
        "phone": {
            "country_code": phone_country_code,
            "number": phone_number,
        },
        "instagram_username": get("instagram_username") or None,
        "facebook_user_id": get("facebook_user_id") or None,
    }


def _row_to_contact_data(row: list[str], field_indices: dict[str, int]) -> dict | None:
    def get(field: str) -> str:
        index = field_indices.get(field)
        if index is None or index >= len(row):
            return ""
        return _cell_value(row[index])

    name = get("name")
    phone_country_code = get("phone_country_code")
    phone_number = get("phone_number")
    if not name or not phone_country_code or not phone_number:
        return None

    return {
        "name": name,
        "phone": {
            "country_code": phone_country_code,
            "number": phone_number,
        },
        "instagram_username": get("instagram_username") or None,
        "facebook_user_id": get("facebook_user_id") or None,
    }


def _map_headers(headers: list[str]) -> dict[str, int]:
    normalized = {_normalize_header(header): index for index, header in enumerate(headers)}
    field_indices: dict[str, int] = {}

    for field, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                field_indices[field] = normalized[alias]
                break

    missing_fields = {"name", "phone_country_code", "phone_number"} - field_indices.keys()
    if missing_fields:
        missing = ", ".join(sorted(missing_fields))
        raise ValueError(f"Missing required columns: {missing}")

    return field_indices


def _iter_csv_contact_rows(file: BinaryIO) -> Iterator[dict]:
    # csv.DictReader reads from the file stream and yields one CSV record at a time.
    text_file = io.TextIOWrapper(file, encoding="utf-8-sig", newline="")
    reader = csv.DictReader(text_file)
    if not reader.fieldnames:
        return

    column_map = _build_column_map(list(reader.fieldnames))
    for row in reader:
        contact_data = _dict_row_to_contact_data(row, column_map)
        if contact_data is not None:
            yield contact_data


def _iter_xlsx_contact_rows(file: BinaryIO) -> Iterator[dict]:
    workbook = load_workbook(file, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            return

        header_row = [_cell_value(header) for header in headers]
        field_indices = _map_headers(header_row)

        for row in rows:
            if row is None or not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_values = [_cell_value(cell) for cell in row]
            contact_data = _row_to_contact_data(row_values, field_indices)
            if contact_data is not None:
                yield contact_data
    finally:
        workbook.close()


def _iter_xls_contact_rows(file: BinaryIO) -> Iterator[dict]:
    content = file.read()
    workbook = xlrd.open_workbook(file_contents=content)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows == 0:
        return

    headers = [_cell_value(sheet.cell_value(0, column)) for column in range(sheet.ncols)]
    field_indices = _map_headers(headers)

    for row_index in range(1, sheet.nrows):
        row_values = [
            _cell_value(sheet.cell_value(row_index, column)) for column in range(sheet.ncols)
        ]
        if not any(value for value in row_values):
            continue
        contact_data = _row_to_contact_data(row_values, field_indices)
        if contact_data is not None:
            yield contact_data


def iter_contact_rows(file: BinaryIO, filename: str) -> Iterator[dict]:
    """Stream contact rows from a file without loading all rows into memory.

    CSV uses the stdlib csv.DictReader, which reads from the file incrementally
    and yields one full CSV record at a time (handles quoted newlines correctly).
    """
    extension = Path(filename).suffix.lower()

    if extension == ".csv":
        yield from _iter_csv_contact_rows(file)
        return
    if extension == ".xlsx":
        yield from _iter_xlsx_contact_rows(file)
        return
    if extension == ".xls":
        yield from _iter_xls_contact_rows(file)
        return

    raise ValueError("Only CSV and Excel files are supported")


def parse_headers(filename: str, content: bytes) -> list[str]:
    extension = Path(filename).suffix.lower()

    if extension == ".csv":
        reader = csv.reader(io.StringIO(content.decode("utf-8-sig")))
        row = next(reader, None)
        return [cell.strip() for cell in row] if row else []
    if extension == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        try:
            row = next(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True), None)
            return [_cell_value(cell) for cell in row] if row else []
        finally:
            workbook.close()
    if extension == ".xls":
        sheet = xlrd.open_workbook(file_contents=content).sheet_by_index(0)
        if sheet.nrows == 0:
            return []
        return [_cell_value(sheet.cell_value(0, column)) for column in range(sheet.ncols)]

    raise ValueError("Only CSV and Excel files are supported")
